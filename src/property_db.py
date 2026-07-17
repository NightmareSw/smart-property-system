"""
property_db.py —— 物业数据库层
=============================
双存储架构：
- SQLite:    结构化数据（住户、公告、物业费），支持精确 CRUD
- Chroma:    公告向量库，支持语义搜索（RAG）

公告的增删改操作会同步双写 SQLite + Chroma，保证数据一致性。
"""

import os
import sys
import sqlite3
import shutil
import hashlib
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from langchain_community.embeddings import SentenceTransformerEmbeddings
from langchain_community.vectorstores import Chroma
from langchain_core.documents import Document

from src.resilience import retry_db

DB_PATH = str(Path(__file__).parent.parent / "property.db")
CHROMA_DIR = str(Path(__file__).parent.parent / "chroma_announcements_db")
ADVER_DIR = str(Path(__file__).parent.parent / "ADVER")
OWNER_EXCEL = str(Path(__file__).parent.parent / "OWNER" / "业主信息.xlsx")
PAYMENT_EXCEL = str(Path(__file__).parent.parent / "OWNER" / "物业费.xlsx")
ADMIN_EXCEL = str(Path(__file__).parent.parent / "admin" / "管理员.xlsx")

# --- 数据库连接管理 ---

def _get_conn() -> sqlite3.Connection:
    """获取SQLite连接（WAL模式，支持并发读写）"""
    conn = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA cache_size=-8000")  # 8MB cache
    return conn


# --- Embedding 模型（全局单例，所有公告共享一个向量库）---
_embed_model = SentenceTransformerEmbeddings(
    model_name=r"C:\Users\Alienware\.cache\modelscope\models\Qwen--Qwen3-Embedding-0.6B\snapshots\master",
    model_kwargs={"device": "cpu"},
    encode_kwargs={"normalize_embeddings": True},
)

# Chroma 向量库实例（模块级单例）
_announcement_vdb: Chroma = None


# ============================================================
# 数据库初始化
# ============================================================

def _hash(plain: str) -> str:
    """简单密码哈希（仅 demo 使用，生产环境请用 bcrypt）"""
    return hashlib.sha256(plain.encode()).hexdigest()


# ============================================================
# 公告向量库（Chroma RAG）—— 语义搜索层
# ============================================================

def _build_announcement_document(ann_id: int, title: str, content: str,
                                  publish_date: str, author: str) -> Document:
    """
    将一条公告转为 Chroma Document。
    文本格式: 标题 + 换行 + 内容，确保标题关键词也被向量化。
    metadata 中保存 SQLite ID，用于删除时定位。
    """
    return Document(
        page_content=f"{title}\n{content}",
        metadata={
            "ann_id": str(ann_id),  # Chroma where 过滤需字符串类型
            "title": title,
            "publish_date": publish_date,
            "author": author,
        }
    )


def _init_chroma():
    """初始化公告 Chroma 向量库（如果已持久化则加载，否则创建）"""
    global _announcement_vdb
    if os.path.exists(CHROMA_DIR) and os.listdir(CHROMA_DIR):
        _announcement_vdb = Chroma(
            persist_directory=CHROMA_DIR,
            embedding_function=_embed_model,
        )
        print("[property_db] 公告向量库已从磁盘加载")
    else:
        _announcement_vdb = Chroma(
            embedding_function=_embed_model,
            persist_directory=CHROMA_DIR,
        )
        print("[property_db] 公告向量库已创建")


def _load_announcements_from_adver() -> list[tuple]:
    """
    从 ADVER 文件夹扫描所有 TXT 文件，每份文件 = 一条公告。

    文件格式（约定）：
      第一行 → 公告标题
      剩余行 → 公告正文
    发布日期取文件的修改时间，作者默认"物业中心"。

    返回: [(title, content, publish_date, author), ...]
    """
    adver_path = Path(ADVER_DIR)
    if not adver_path.exists():
        return []

    results = []
    # 扫描所有 TXT，按文件名排序保证稳定
    for filepath in sorted(adver_path.glob("*.txt")):
        try:
            for encoding in ['utf-8', 'gbk']:
                try:
                    with open(filepath, 'r', encoding=encoding) as f:
                        lines = [l.strip() for l in f.readlines() if l.strip()]
                    if len(lines) >= 2:
                        title = lines[0]
                        content = "\n".join(lines[1:])
                    elif len(lines) == 1:
                        title = lines[0]
                        content = lines[0]
                    else:
                        continue

                    pub_date = datetime.fromtimestamp(
                        filepath.stat().st_mtime
                    ).strftime("%Y-%m-%d")
                    results.append((title, content, pub_date, "物业中心"))
                    break
                except UnicodeDecodeError:
                    continue
        except Exception as e:
            print(f"[property_db] 读取公告文件失败 {filepath.name}: {e}")

    return results


def _seed_chroma():
    """将 SQLite 中已有的公告写入 Chroma（仅当 Chroma 为空时执行）"""
    # 检查 Chroma 是否已有数据，避免重复导入导致向量膨胀
    if _announcement_vdb._collection.count() > 0:
        return

    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT id, title, content, publish_date, author FROM announcements")
    rows = cur.fetchall()
    conn.close()

    if not rows:
        return

    docs = [
        _build_announcement_document(r["id"], r["title"], r["content"],
                                     r["publish_date"], r["author"])
        for r in rows
    ]
    _announcement_vdb.add_documents(docs)
    _announcement_vdb.persist()
    print(f"[property_db] Chroma 种子数据写入完成: {len(docs)} 条公告")


def _chroma_add_announcement(ann_id: int, title: str, content: str,
                              publish_date: str, author: str):
    """向 Chroma 新增一条公告（与 SQLite INSERT 同步调用）"""
    doc = _build_announcement_document(ann_id, title, content, publish_date, author)
    _announcement_vdb.add_documents([doc])
    _announcement_vdb.persist()


def _chroma_delete_announcement(ann_id: int):
    """从 Chroma 删除一条公告（通过 metadata.ann_id 过滤）"""
    try:
        collection = _announcement_vdb._collection
        collection.delete(where={"ann_id": str(ann_id)})
        _announcement_vdb.persist()
    except Exception as e:
        print(f"[property_db] Chroma 删除失败 (id={ann_id}): {e}")


def _chroma_update_announcement(ann_id: int, title: str, content: str,
                                 publish_date: str, author: str):
    """更新 Chroma 中的公告：先删旧文档，再插入新文档"""
    _chroma_delete_announcement(ann_id)
    _chroma_add_announcement(ann_id, title, content, publish_date, author)


def search_announcements_rag(query: str, k: int = 4) -> list[dict]:
    """
    RAG 语义搜索公告 —— 用自然语言查询相关公告。

    与 SQLite LIKE 不同，这里不要求关键词精确匹配。
    "最近有什么安全方面的通知？" 能匹配到"电梯维护公告"。

    Args:
        query: 自然语言查询
        k: 返回最相似的 k 条结果

    Returns:
        公告列表（dict 格式，含相似度分数）
    """
    if _announcement_vdb is None:
        return []

    docs = _announcement_vdb.similarity_search_with_score(query, k=k)
    results = []
    for doc, score in docs:
        results.append({
            "id": int(doc.metadata.get("ann_id", 0)),
            "title": doc.metadata.get("title", ""),
            "content": doc.page_content,
            "publish_date": doc.metadata.get("publish_date", ""),
            "author": doc.metadata.get("author", ""),
            "score": round(float(1.0 - score), 4),  # distance → similarity
        })
    return results


# ============================================================
# ============================================================
# Excel 种子数据读取（仅首次初始化时使用）
# ============================================================

def _read_owners_excel() -> list[dict]:
    """从 OWNER/业主信息.xlsx 读取业主数据（仅用于首次导入 SQLite）"""
    import openpyxl
    path = Path(OWNER_EXCEL)
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    owners = []
    for row in rows:
        if not row[0]:
            continue
        owners.append({
            "room_number": str(row[0]).strip(),
            "password": str(row[1]).strip() if row[1] else "",
            "owner_name": str(row[2]).strip() if row[2] else "",
            "phone": str(row[3]).strip() if len(row) > 3 and row[3] else "",
        })
    return owners


def _read_payments_excel() -> list[tuple]:
    """从 OWNER/物业费.xlsx 读取物业费数据（仅用于首次导入 SQLite）"""
    import openpyxl
    path = Path(PAYMENT_EXCEL)
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    result = []
    for row in rows:
        if not row[0]:
            continue
        result.append((
            str(row[0]).strip(),
            float(row[1]) if row[1] else 0.0,
            str(row[2]).strip() if row[2] else "",
            str(row[3]).strip() if row[3] else "unpaid",
            str(row[4]).strip() if len(row) > 4 and row[4] else "",
        ))
    return result


def _read_admins_excel() -> list[dict]:
    """从 admin/管理员.xlsx 读取管理员数据（仅用于首次导入 SQLite）"""
    import openpyxl
    path = Path(ADMIN_EXCEL)
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    admins = []
    for row in rows:
        if not row[0]:
            continue
        admins.append({
            "username": str(row[0]).strip(),
            "password": str(row[1]).strip() if row[1] else "",
            "name": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            "role": str(row[3]).strip() if len(row) > 3 and row[3] else "admin",
        })
    return admins


# ============================================================
# 数据库初始化
# ============================================================

def init_db():
    """
    初始化数据库：
    - SQLite: announcements / owners / payments / admins 四张表
    - 首次启动时从 Excel 导入种子数据
    - Chroma: 公告向量库（RAG 语义搜索）
    """
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # --- 建表 ---
    cur.execute("""
        CREATE TABLE IF NOT EXISTS announcements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            content TEXT NOT NULL,
            publish_date TEXT NOT NULL,
            author TEXT DEFAULT '物业中心'
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS owners (
            room_number TEXT PRIMARY KEY,
            password TEXT NOT NULL,
            owner_name TEXT NOT NULL,
            phone TEXT DEFAULT ''
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            room_number TEXT NOT NULL,
            amount REAL NOT NULL,
            due_date TEXT NOT NULL,
            status TEXT DEFAULT 'unpaid',
            paid_date TEXT DEFAULT '',
            notes TEXT DEFAULT ''
        )
    """)

    # --- 迁移：添加 notes 列 ---
    try:
        cur.execute("ALTER TABLE payments ADD COLUMN notes TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass

    cur.execute("""
        CREATE TABLE IF NOT EXISTS admins (
            username TEXT PRIMARY KEY,
            password TEXT NOT NULL,
            name TEXT DEFAULT '',
            role TEXT DEFAULT 'admin'
        )
    """)

    # --- 迁移：添加 source_file 列 ---
    try:
        cur.execute("ALTER TABLE announcements ADD COLUMN source_file TEXT")
    except sqlite3.OperationalError:
        pass

    # --- 公告种子数据（从 ADVER/）---
    cur.execute("SELECT COUNT(*) FROM announcements")
    if cur.fetchone()[0] == 0:
        announcements = _load_announcements_from_adver()
        if announcements:
            cur.executemany(
                "INSERT INTO announcements (title, content, publish_date, author) "
                "VALUES (?,?,?,?)", announcements)
            print(f"[property_db] 从 ADVER/ 导入了 {len(announcements)} 条公告")

    # --- 业主种子数据（从 OWNER/业主信息.xlsx）---
    cur.execute("SELECT COUNT(*) FROM owners")
    if cur.fetchone()[0] == 0:
        owners = _read_owners_excel()
        if owners:
            cur.executemany(
                "INSERT INTO owners (room_number, password, owner_name, phone) "
                "VALUES (?,?,?,?)",
                [(o["room_number"], o["password"], o["owner_name"], o["phone"])
                 for o in owners])
            print(f"[property_db] 从 Excel 导入了 {len(owners)} 户业主")

    # --- 物业费种子数据（从 OWNER/物业费.xlsx）---
    cur.execute("SELECT COUNT(*) FROM payments")
    if cur.fetchone()[0] == 0:
        payments = _read_payments_excel()
        if payments:
            cur.executemany(
                "INSERT INTO payments (room_number, amount, due_date, status, paid_date, notes) "
                "VALUES (?,?,?,?,?,'')", payments)
            print(f"[property_db] 从 Excel 导入了 {len(payments)} 条物业费记录")

    # --- 管理员种子数据（从 admin/管理员.xlsx）---
    cur.execute("SELECT COUNT(*) FROM admins")
    if cur.fetchone()[0] == 0:
        admins = _read_admins_excel()
        if admins:
            cur.executemany(
                "INSERT INTO admins (username, password, name, role) VALUES (?,?,?,?)",
                [(a["username"], a["password"], a["name"], a["role"]) for a in admins])
            print(f"[property_db] 从 Excel 导入了 {len(admins)} 个管理员账号")

    # --- 通知表 ---
    cur.execute("""
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            room_number TEXT DEFAULT '',
            title TEXT NOT NULL,
            content TEXT NOT NULL,
            type TEXT DEFAULT 'info',
            is_read INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        )
    """)

    # --- 报修工单表 ---
    cur.execute("""
        CREATE TABLE IF NOT EXISTS repairs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            room_number TEXT NOT NULL,
            title TEXT NOT NULL,
            description TEXT NOT NULL,
            status TEXT DEFAULT 'pending',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            admin_note TEXT DEFAULT ''
        )
    """)

    # --- 情感分析日志表 ---
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sentiment_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            room_number TEXT DEFAULT '',
            message TEXT NOT NULL,
            score REAL DEFAULT 0.5,
            label TEXT DEFAULT 'neutral',
            entropy_high INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        )
    """)

    # --- 迁移：添加 entropy_high 列 ---
    try:
        cur.execute("ALTER TABLE sentiment_logs ADD COLUMN entropy_high INTEGER DEFAULT 0")
    except sqlite3.OperationalError:
        pass

    # --- 物业费操作日志表 ---
    cur.execute("""
        CREATE TABLE IF NOT EXISTS payment_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT NOT NULL,
            admin_name TEXT DEFAULT '',
            room_count INTEGER DEFAULT 0,
            amount REAL DEFAULT 0,
            period_text TEXT DEFAULT '',
            payment_ids TEXT DEFAULT '',
            created_at TEXT NOT NULL
        )
    """)

    conn.commit()
    conn.close()
    print("[property_db] SQLite 初始化完成")

    # --- 初始化公告向量库 ---
    _init_chroma()
    # 首次运行时 SQLite demo 数据刚插入，同步到 Chroma
    _seed_chroma()


# ============================================================
# 业主 CRUD —— SQLite
# ============================================================

def get_all_owners() -> list[dict]:
    """查询全部业主"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT room_number, password, owner_name, phone FROM owners ORDER BY room_number")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def add_owner(room_number: str, password: str, owner_name: str, phone: str = "") -> bool:
    """新增业主，门牌号已存在时返回 False"""
    try:
        conn = _get_conn()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO owners (room_number, password, owner_name, phone) VALUES (?,?,?,?)",
            (str(room_number).strip(), password, owner_name, phone))
        conn.commit()
        conn.close()
        return True
    except sqlite3.IntegrityError:
        return False


def update_owner(room_number: str, owner_name: str = None, phone: str = None,
                 password: str = None) -> bool:
    """更新业主信息，未传的字段保持不变"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM owners WHERE room_number=?", (str(room_number).strip(),))
    row = cur.fetchone()
    if not row:
        conn.close()
        return False
    new_name = owner_name if owner_name is not None else row["owner_name"]
    new_phone = phone if phone is not None else row["phone"]
    new_pwd = password if password is not None else row["password"]
    cur.execute(
        "UPDATE owners SET owner_name=?, phone=?, password=? WHERE room_number=?",
        (new_name, new_phone, new_pwd, str(room_number).strip()))
    conn.commit()
    conn.close()
    return True


def delete_owner(room_number: str) -> bool:
    """删除业主"""
    conn = _get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM owners WHERE room_number=?", (str(room_number).strip(),))
    deleted = cur.rowcount > 0
    conn.commit()
    conn.close()
    return deleted


# ============================================================
# 认证
# ============================================================

def authenticate(room_number: str, password: str) -> dict | None:
    """验证住户登录"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        "SELECT room_number, owner_name, phone FROM owners WHERE room_number=? AND password=?",
        (str(room_number).strip(), password))
    row = cur.fetchone()
    conn.close()
    return {"room_number": row["room_number"], "owner_name": row["owner_name"],
            "phone": row["phone"]} if row else None


# ============================================================
# 管理员账号 CRUD —— SQLite
# ============================================================

def get_all_admins() -> list[dict]:
    """查询全部管理员"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT username, password, name, role FROM admins")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def authenticate_admin(username: str, password: str) -> dict | None:
    """验证管理员登录"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        "SELECT username, name, role FROM admins WHERE username=? AND password=?",
        (username, password))
    row = cur.fetchone()
    conn.close()
    return {"username": row["username"], "name": row["name"],
            "role": row["role"]} if row else None


def add_admin(username: str, password: str, name: str = "", role: str = "admin") -> bool:
    """新增管理员，用户名已存在时返回 False"""
    try:
        conn = _get_conn()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO admins (username, password, name, role) VALUES (?,?,?,?)",
            (username, password, name, role))
        conn.commit()
        conn.close()
        return True
    except sqlite3.IntegrityError:
        return False


def update_admin(username: str, password: str = None, name: str = None,
                 role: str = None) -> bool:
    """更新管理员信息"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM admins WHERE username=?", (username,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return False
    new_pwd = password if password is not None else row["password"]
    new_name = name if name is not None else row["name"]
    new_role = role if role is not None else row["role"]
    cur.execute(
        "UPDATE admins SET password=?, name=?, role=? WHERE username=?",
        (new_pwd, new_name, new_role, username))
    conn.commit()
    conn.close()
    return True


def delete_admin(username: str) -> bool:
    """删除管理员"""
    conn = _get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM admins WHERE username=?", (username,))
    deleted = cur.rowcount > 0
    conn.commit()
    conn.close()
    return deleted


# ============================================================
# 公告 CRUD（AdminAgent 全权限，ResidentAgent 只读）
# ============================================================

def get_all_announcements() -> list[dict]:
    """查询全部公告"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT id, title, content, publish_date, author FROM announcements ORDER BY id DESC")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def search_announcements(keyword: str) -> list[dict]:
    """按关键词搜索公告（模糊匹配标题和内容）"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        "SELECT id, title, content, publish_date, author FROM announcements "
        "WHERE title LIKE ? OR content LIKE ? ORDER BY publish_date DESC",
        (f"%{keyword}%", f"%{keyword}%")
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


@retry_db(max_retries=3)
def add_announcement(title: str, content: str, author: str = "物业中心",
                     source_file: str = None) -> int:
    """新增公告（SQLite + Chroma 双写），返回新记录的 id"""
    pub_date = datetime.now().strftime("%Y-%m-%d")
    conn = _get_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO announcements (title, content, publish_date, author, source_file) "
        "VALUES (?,?,?,?,?)",
        (title, content, pub_date, author, source_file)
    )
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    _chroma_add_announcement(new_id, title, content, pub_date, author)
    return new_id


def update_announcement(ann_id: int, title: str = None, content: str = None) -> bool:
    """更新公告标题和/或内容（SQLite + Chroma 双更新）"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM announcements WHERE id=?", (ann_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return False

    new_title = title if title else row["title"]
    new_content = content if content else row["content"]

    if title:
        cur.execute("UPDATE announcements SET title=? WHERE id=?", (title, ann_id))
    if content:
        cur.execute("UPDATE announcements SET content=? WHERE id=?", (content, ann_id))
    conn.commit()
    conn.close()

    # 同步更新 Chroma 向量库
    _chroma_update_announcement(ann_id, new_title, new_content,
                                row["publish_date"], row["author"])
    return True


def delete_announcement(ann_id: int) -> bool:
    """删除公告（SQLite + Chroma 双删）"""
    conn = _get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM announcements WHERE id=?", (ann_id,))
    if not cur.fetchone():
        conn.close()
        return False
    cur.execute("DELETE FROM announcements WHERE id=?", (ann_id,))
    conn.commit()
    conn.close()
    # 同步从 Chroma 向量库删除
    _chroma_delete_announcement(ann_id)
    return True


def delete_announcement_by_source(source_file: str) -> bool:
    """按 ADVER 源文件名删除公告（SQLite + Chroma 双删）"""
    conn = _get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM announcements WHERE source_file=?", (source_file,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return False
    ann_id = row[0]
    cur.execute("DELETE FROM announcements WHERE id=?", (ann_id,))
    conn.commit()
    conn.close()
    _chroma_delete_announcement(ann_id)
    return True


def delete_all_announcements():
    """清空全部公告（SQLite + Chroma），用于重建索引"""
    conn = _get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM announcements")
    conn.commit()
    conn.close()
    try:
        _announcement_vdb._collection.delete(where={})
        _announcement_vdb.persist()
        print("[property_db] Chroma 公告集合已清空")
    except Exception as e:
        print(f"[property_db] Chroma 清空失败: {e}")


def parse_uploaded_file(filepath: str) -> tuple:
    """
    解析上传的文件（TXT 或 PDF），返回 (title, content)。

    TXT: 第一非空行 = 标题，剩余行 = 正文
    PDF: 文件名（无扩展名）= 标题，全部提取文本 = 正文
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".txt":
        for encoding in ["utf-8", "gbk"]:
            try:
                with open(filepath, "r", encoding=encoding) as f:
                    lines = [l.strip() for l in f.readlines() if l.strip()]
                if len(lines) >= 2:
                    return lines[0], "\n".join(lines[1:])
                elif len(lines) == 1:
                    return lines[0], lines[0]
                else:
                    raise ValueError("文件内容为空")
            except UnicodeDecodeError:
                continue
        raise ValueError("无法解码TXT文件")

    elif ext == ".pdf":
        from langchain_community.document_loaders import PyPDFLoader
        loader = PyPDFLoader(filepath)
        pages = loader.load()
        if not pages:
            raise ValueError("PDF文件为空或无法解析")
        title = os.path.splitext(os.path.basename(filepath))[0]
        content = ""
        for i, page in enumerate(pages):
            content += f"\n--- 第{i+1}页 ---\n{page.page_content}"
        return title, content.strip()

    else:
        raise ValueError(f"不支持的文件类型: {ext}")


# ============================================================
# 物业费 CRUD —— SQLite
# ============================================================

def get_all_payments() -> list[dict]:
    """查询全部物业费记录"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM payments ORDER BY id")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def get_payment_by_room(room_number: str) -> list[dict]:
    """按门牌号查询物业费"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM payments WHERE room_number=? ORDER BY id",
                (str(room_number).strip(),))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def get_payments_by_status(status: str) -> list[dict]:
    """按缴费状态筛选物业费记录"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM payments WHERE status=? ORDER BY id", (status,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def get_db_stats() -> dict:
    """获取数据库统计信息"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM announcements")
    ann_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM owners")
    owner_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM payments")
    pay_count = cur.fetchone()[0]

    cur.execute("SELECT status, COUNT(*) FROM payments GROUP BY status")
    status_counts = {row["status"]: row["COUNT(*)"] for row in cur.fetchall()}

    conn.close()

    chroma_count = 0
    try:
        if _announcement_vdb:
            chroma_count = _announcement_vdb._collection.count()
    except Exception:
        pass

    return {
        "announcements": ann_count,
        "chroma_vectors": chroma_count,
        "payments": pay_count,
        "paid": status_counts.get("paid", 0),
        "unpaid": status_counts.get("unpaid", 0),
        "overdue": status_counts.get("overdue", 0),
        "owners": owner_count,
    }


@retry_db(max_retries=3)
def add_payment(room_number: str, amount: float, due_date: str,
                status: str = "unpaid", notes: str = "") -> int:
    """新增物业费记录，返回新记录的 id"""
    conn = _get_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO payments (room_number, amount, due_date, status, notes) "
        "VALUES (?,?,?,?,?)",
        (str(room_number).strip(), amount, due_date, status, notes))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return new_id


def update_payment_status(pay_id: int, status: str, paid_date: str = None) -> bool:
    """更新缴费状态"""
    conn = _get_conn()
    cur = conn.cursor()
    pd = paid_date or datetime.now().strftime("%Y-%m-%d")
    cur.execute("UPDATE payments SET status=?, paid_date=? WHERE id=?",
                (status, pd, pay_id))
    updated = cur.rowcount > 0
    conn.commit()
    conn.close()
    return updated


def delete_payment(pay_id: int) -> bool:
    """删除物业费记录"""
    conn = _get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM payments WHERE id=?", (pay_id,))
    deleted = cur.rowcount > 0
    conn.commit()
    conn.close()
    return deleted


# ============================================================
# 消息通知
# ============================================================

def add_notification(room_number: str, title: str, content: str, ntype: str = "info") -> int:
    """新增通知，room_number 为空表示全员通知"""
    conn = _get_conn()
    cur = conn.cursor()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        "INSERT INTO notifications (room_number, title, content, type, created_at) "
        "VALUES (?,?,?,?,?)",
        (room_number, title, content, ntype, now))
    conn.commit()
    nid = cur.lastrowid
    conn.close()
    return nid


def get_notifications(room_number: str = "", include_public: bool = True) -> list[dict]:
    """获取通知列表。include_public=False 时仅显示该住户专属通知（隐私保护）"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    if include_public:
        cur.execute(
            "SELECT * FROM notifications WHERE room_number=? OR room_number='' "
            "ORDER BY id DESC LIMIT 50", (room_number,))
    else:
        cur.execute(
            "SELECT * FROM notifications WHERE room_number=? "
            "ORDER BY id DESC LIMIT 50", (room_number,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def get_unread_count(room_number: str, include_public: bool = True) -> int:
    """获取未读通知数量"""
    conn = _get_conn()
    cur = conn.cursor()
    if include_public:
        cur.execute(
            "SELECT COUNT(*) FROM notifications "
            "WHERE is_read=0 AND (room_number=? OR room_number='')", (room_number,))
    else:
        cur.execute(
            "SELECT COUNT(*) FROM notifications "
            "WHERE is_read=0 AND room_number=?", (room_number,))
    count = cur.fetchone()[0]
    conn.close()
    return count


def mark_notification_read(nid: int) -> bool:
    """标记通知为已读"""
    conn = _get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE notifications SET is_read=1 WHERE id=?", (nid,))
    updated = cur.rowcount > 0
    conn.commit()
    conn.close()
    return updated


# ============================================================
# 报修工单
# ============================================================

@retry_db(max_retries=3)
def add_repair(room_number: str, title: str, description: str) -> int:
    """住户提交报修，返回工单ID"""
    conn = _get_conn()
    cur = conn.cursor()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        "INSERT INTO repairs (room_number, title, description, status, created_at, updated_at) "
        "VALUES (?,?,?,'pending',?,?)",
        (room_number, title, description, now, now))
    conn.commit()
    rid = cur.lastrowid
    conn.close()
    return rid


def get_all_repairs() -> list[dict]:
    """管理员查看全部工单"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM repairs ORDER BY id DESC")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def get_repairs_by_room(room_number: str) -> list[dict]:
    """住户查看自己的工单"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM repairs WHERE room_number=? ORDER BY id DESC",
                (str(room_number).strip(),))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def update_repair(repair_id: int, status: str = None, admin_note: str = None) -> bool:
    """管理员更新工单状态/备注"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM repairs WHERE id=?", (repair_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return False
    new_status = status if status else row["status"]
    new_note = admin_note if admin_note is not None else row["admin_note"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        "UPDATE repairs SET status=?, admin_note=?, updated_at=? WHERE id=?",
        (new_status, new_note, now, repair_id))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 物业费操作日志与撤销
# ============================================================

def add_payment_log(action: str, admin_name: str, room_count: int, amount: float,
                    period_text: str, payment_ids: list[int]) -> int:
    """记录物业费发布/撤销操作，返回日志ID"""
    conn = _get_conn()
    cur = conn.cursor()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        "INSERT INTO payment_logs (action, admin_name, room_count, amount, "
        "period_text, payment_ids, created_at) VALUES (?,?,?,?,?,?,?)",
        (action, admin_name, room_count, amount, period_text,
         ",".join(str(i) for i in payment_ids), now))
    conn.commit()
    lid = cur.lastrowid
    conn.close()
    return lid


def get_payment_logs(limit: int = 20) -> list[dict]:
    """获取最近的物业费操作日志"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM payment_logs ORDER BY id DESC LIMIT ?", (limit,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def undo_payment_log(log_id: int) -> dict:
    """撤销指定日志对应的物业费发布操作，返回撤销信息"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT * FROM payment_logs WHERE id=?", (log_id,))
    log = cur.fetchone()
    if not log:
        conn.close()
        return None
    if log["action"] != "publish":
        conn.close()
        return {"error": "仅可撤销发布操作"}

    ids = [int(i) for i in log["payment_ids"].split(",") if i]
    deleted = 0
    for pid in ids:
        cur.execute("DELETE FROM payments WHERE id=?", (pid,))
        deleted += cur.rowcount

    # 将日志标记为已撤销
    cur.execute("UPDATE payment_logs SET action='undone' WHERE id=?", (log_id,))
    conn.commit()
    conn.close()

    return {
        "log_id": log_id,
        "deleted": deleted,
        "period": log["period_text"],
        "room_count": log["room_count"],
    }


# ============================================================
# 情感分析日志
# ============================================================

def add_sentiment_log(room_number: str, message: str, score: float, label: str,
                      entropy_high: bool = False) -> int:
    """记录一条情感分析结果，可选标记是否触发了信息熵反问"""
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO sentiment_logs (room_number, message, score, label, entropy_high, created_at) "
        "VALUES (?,?,?,?,?,?)",
        (room_number, message, round(score, 4), label, int(entropy_high),
         datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    sid = cur.lastrowid
    conn.commit()
    conn.close()
    return sid


def get_sentiment_stats(hours: int = 24) -> dict:
    """
    获取情感分析统计数据。
    hours: 统计最近多少小时的数据（默认24小时）
    返回: {total, positive, neutral, negative, avg_score, timeline}
    """
    conn = _get_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    since = (datetime.now().strftime("%Y-%m-%d %H:%M:%S")
             if hours <= 0 else
             "")
    if hours > 0:
        from datetime import timedelta
        since = (datetime.now() - timedelta(hours=hours)).strftime("%Y-%m-%d %H:%M:%S")

    cur.execute("SELECT COUNT(*) as cnt FROM sentiment_logs WHERE created_at >= ?", (since,))
    total = cur.fetchone()["cnt"]

    if total == 0:
        conn.close()
        return {"total": 0, "positive": 0, "neutral": 0, "negative": 0, "avg_score": 0, "recent": []}

    cur.execute(
        "SELECT label, COUNT(*) as cnt FROM sentiment_logs "
        "WHERE created_at >= ? GROUP BY label", (since,))
    label_counts = {"positive": 0, "neutral": 0, "negative": 0}
    for row in cur.fetchall():
        label_counts[row["label"]] = row["cnt"]

    cur.execute(
        "SELECT AVG(score) as avg_score FROM sentiment_logs WHERE created_at >= ?", (since,))

    avg_score = cur.fetchone()["avg_score"] or 0

    cur.execute(
        "SELECT room_number, score, label, message, entropy_high, created_at "
        "FROM sentiment_logs WHERE created_at >= ? AND label = 'negative' "
        "ORDER BY created_at DESC LIMIT 10", (since,))
    recent = [dict(r) for r in cur.fetchall()]

    cur.execute(
        "SELECT COUNT(*) as cnt FROM sentiment_logs "
        "WHERE created_at >= ? AND entropy_high = 1", (since,))
    entropy_count = cur.fetchone()["cnt"]

    conn.close()
    return {
        "total": total,
        "positive": label_counts["positive"],
        "neutral": label_counts["neutral"],
        "negative": label_counts["negative"],
        "avg_score": round(avg_score, 4),
        "entropy_triggers": entropy_count,
        "recent": recent,
    }


# ============================================================
# 启动时自动初始化（import 即执行）
# ============================================================
if __name__ == "__main__":
    # 直接运行时重建：清理旧数据，重新初始化
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    if os.path.exists(CHROMA_DIR):
        shutil.rmtree(CHROMA_DIR)
    print("旧数据已清理，重新初始化...")
    init_db()
    print("DB created & seeded.")
else:
    # import 时只初始化一次（SQLite 建表幂等，Chroma 加载已有数据）
    init_db()
